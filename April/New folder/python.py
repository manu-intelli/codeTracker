import os
import re
import xlrd
import openpyxl
import pandas as pd

# Update this with your folder path
folder_path = "your_folder_path_here"
output_file = "final_summary1.xlsx"

data = []

# Define regex patterns (case-insensitive)
patterns = {
    "Part# / Model name": r"(part\s*#|model\s*name)",
    "OPP#": r"opp\s*#?",
    "CUSTOMER": r"customer",
    "Assembly cost / PPD": r"\b(assembly cost|ppd)\b",
    "Estimated BOM cost": r"\b(estimated bom cost|bom cost per unit)\b",
    "Design & Development cost": r"design and development cost",
    "Recommended Price": r"recommended price",
    "Comments to Steven": r"(comments for steven\.s|comments to steven)",
    "CREATED ON": r"created\s*on\s*[:\-]?"
}

# Force value as string for "CREATED ON"
def clean_value(value, key):
    value = str(value).strip()
    if key == "CREATED ON":
        return "'" + value if not value.startswith("'") else value
    return value

# XLS file reader
def extract_from_xls(sheet):
    extracted = {}
    for row_idx in range(sheet.nrows):
        for col_idx in range(sheet.ncols):
            cell_value = str(sheet.cell_value(row_idx, col_idx)).strip().lower()
            for key, pattern in patterns.items():
                if re.search(pattern, cell_value, re.IGNORECASE):
                    try:
                        next_value = sheet.cell_value(row_idx, col_idx + 1)
                        if key not in extracted:
                            extracted[key] = clean_value(next_value, key)
                    except:
                        continue
    return extracted

# XLSX file reader
def extract_from_xlsx(sheet):
    extracted = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                value = str(cell.value).strip().lower()
                for key, pattern in patterns.items():
                    if re.search(pattern, value, re.IGNORECASE):
                        try:
                            next_cell_value = sheet.cell(cell.row, cell.column + 1).value
                            if key not in extracted:
                                extracted[key] = clean_value(next_cell_value, key)
                        except:
                            continue
    return extracted

# Loop through all Excel files in folder
for filename in os.listdir(folder_path):
    if filename.endswith(".xls") or filename.endswith(".xlsx"):
        file_path = os.path.join(folder_path, filename)
        row_data = {"File Name": filename}

        try:
            if filename.endswith(".xls"):
                book = xlrd.open_workbook(file_path)
                sheet = book.sheet_by_index(0)
                extracted = extract_from_xls(sheet)
            else:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                sheet = wb.active
                extracted = extract_from_xlsx(sheet)

            row_data.update(extracted)
            data.append(row_data)

        except Exception as e:
            print(f"❌ Error reading {filename}: {e}")

# Save to final Excel file with all columns as text
df = pd.DataFrame(data)

with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
    df.to_excel(writer, index=False, sheet_name='Summary')
    workbook = writer.book
    worksheet = writer.sheets['Summary']

    text_format = workbook.add_format({'num_format': '@'})  # text format
    for col_num in range(len(df.columns)):
        worksheet.set_column(col_num, col_num, 25, text_format)

print(f"✅ Summary saved to {output_file}")
