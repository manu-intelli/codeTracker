import os
import re
import xlrd
import openpyxl
import pandas as pd

folder_path = "your_folder_path_here"
output_file = "final_summary1.xlsx"

data = []

# Define regex patterns (case-insensitive)
patterns = {
    "Part# / Model name": r"(part\s*#|model\s*name)",  # Matching part# and model name
    "OPP#": r"opp\s*#?",  # Matching OPP#
    "CUSTOMER": r"customer",
    "Assembly cost / PPD": r"\b(assembly cost|ppd)\b",  # Matching assembly cost or PPD
    "Estimated BOM cost": r"\b(estimated bom cost|bom cost per unit)\b",  # Matching BOM cost
    "Design & Development cost": r"design and development cost",  # Matching design and development cost
    "Recommended Price": r"recommended price",  # Matching recommended price
    "Comments to Steven": r"(comments for steven\.s|comments to steven)",  # Matching comments for Steven
    "CREATED ON": r"created\s*on\s*[:\-]?"  # Matching created on
}

def clean_value(value, key):
    value = str(value).strip()
    # For "CREATED ON", force it to be a string in Excel
    if key == "CREATED ON":
        return "'" + value  # prevent Excel auto-conversion to serial
    return value

def extract_from_xls(sheet):
    extracted = {}
    for row_idx in range(sheet.nrows):
        for col_idx in range(sheet.ncols):
            cell_value = str(sheet.cell_value(row_idx, col_idx)).strip().lower()
            for key, pattern in patterns.items():
                if re.search(pattern, cell_value, re.IGNORECASE):
                    try:
                        next_value = sheet.cell_value(row_idx, col_idx + 1)
                        if key not in extracted:
                            extracted[key] = clean_value(next_value, key)
                    except:
                        continue
    return extracted

def extract_from_xlsx(sheet):
    extracted = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                value = str(cell.value).strip().lower()
                for key, pattern in patterns.items():
                    if re.search(pattern, value, re.IGNORECASE):
                        try:
                            next_cell_value = sheet.cell(cell.row, cell.column + 1).value
                            if key not in extracted:
                                extracted[key] = clean_value(next_cell_value, key)
                        except:
                            continue
    return extracted

# Loop through each file in folder
for filename in os.listdir(folder_path):
    if filename.endswith(".xls") or filename.endswith(".xlsx"):
        file_path = os.path.join(folder_path, filename)
        row_data = {"File Name": filename}

        try:
            if filename.endswith(".xls"):
                book = xlrd.open_workbook(file_path)
                sheet = book.sheet_by_index(0)
                extracted = extract_from_xls(sheet)
            else:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                sheet = wb.active
                extracted = extract_from_xlsx(sheet)

            row_data.update(extracted)
            data.append(row_data)

        except Exception as e:
            print(f"❌ Error reading {filename}: {e}")

# Save to final Excel file
df = pd.DataFrame(data)

# Optional: Force text format using xlsxwriter to avoid Excel formatting issues
with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
    df.to_excel(writer, index=False, sheet_name='Summary')
    workbook = writer.book
    worksheet = writer.sheets['Summary']
    text_format = workbook.add_format({'num_format': '@'})  # Text format
    for col_num in range(len(df.columns)):
        worksheet.set_column(col_num, col_num, 25, text_format)

print(f"✅ Summary saved to {output_file}")