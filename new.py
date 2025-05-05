import os
import re
import xlrd
import openpyxl
from datetime import datetime
from openpyxl.cell.cell import Cell
import pandas as pd

# Main folder path
main_folder_path = r"D:\RIghtstroken\Pricing\PricingExtraction\Pricing_Sheets_24-Feb-25"
output_file = "final_summary_all_folders.xlsx"

# Field patterns
patterns = {
    "Part# / Model name": r"(part\s*#|model\s*name)",
    "OPP#": r"opp\s*#?",
    "CUSTOMER": r"\bcustomer\b|\bcustomer\s*name\b|\bclient\s*name\b",
    "Assembly cost / PPD": r"\b(assembly cost|ppd)\b",
    "Estimated BOM cost": r"\b(estimated bom cost|bom cost per unit)\b",
    "Design & Development cost": r"design and development cost",
    "Recommended Price": r"recommended price",
    "Comments to Steven": r"(comments for steven\.s|comments to steven)",
    "CREATED ON": r"created\s*on\s*[:\-]?"
}

# Clean and extract value, support for in-cell and date formats
def clean_value(value, key, cell=None):
    if isinstance(cell, Cell) and cell.is_date:
        return cell.value.strftime("%m/%d/%Y")

    if key == "CREATED ON":
        # Try converting Excel float date
        if isinstance(value, float):
            try:
                date_value = datetime(*xlrd.xldate_as_tuple(value, 0))
                return date_value.strftime("%m/%d/%Y")
            except:
                pass

        # Try matching string pattern inside same cell
        if isinstance(value, str):
            match = re.search(r"(\d{1,2})[-/\s](\w{3,})[-/\s](\d{2,4})", value, re.IGNORECASE)
            if match:
                for fmt in ["%d-%b-%y", "%d-%b-%Y", "%d/%b/%y", "%d/%b/%Y"]:
                    try:
                        return datetime.strptime(match.group(0), fmt).strftime("%m/%d/%Y")
                    except:
                        continue

    return str(value).strip()

# Extract from XLS
def extract_from_xls(sheet):
    extracted = {}
    for row_idx in range(sheet.nrows):
        for col_idx in range(sheet.ncols):
            cell_text = str(sheet.cell_value(row_idx, col_idx)).strip().lower()
            for key, pattern in patterns.items():
                if key not in extracted and re.search(pattern, cell_text, re.IGNORECASE):
                    # Try in same cell
                    cleaned = clean_value(sheet.cell_value(row_idx, col_idx), key)
                    if cleaned and cleaned.lower() != cell_text:
                        extracted[key] = cleaned
                        continue

                    # Try next cell
                    next_val = ""
                    try:
                        if col_idx + 1 < sheet.ncols:
                            next_val = sheet.cell_value(row_idx, col_idx + 1)
                        elif row_idx + 1 < sheet.nrows:
                            next_val = sheet.cell_value(row_idx + 1, col_idx)
                    except:
                        pass

                    extracted[key] = clean_value(next_val, key)
    return extracted

# Extract from XLSX
def extract_from_xlsx(sheet):
    extracted = {}
    max_row, max_col = sheet.max_row, sheet.max_column

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                cell_text = str(cell.value).strip().lower()
                for key, pattern in patterns.items():
                    if key not in extracted or not extracted[key]:  # Only if key is missing or empty
                        if re.search(pattern, cell_text, re.IGNORECASE):
                            # Debugging print to trace matches
                            print(f"Found '{key}' match in cell: {cell_text}")

                            # First check value in same cell
                            cleaned = clean_value(cell.value, key, cell)
                            if cleaned and cleaned.lower() != cell_text:
                                extracted[key] = cleaned
                                continue

                            # Try next cells
                            next_val = None
                            try:
                                if cell.column + 1 <= max_col:
                                    next_val = sheet.cell(cell.row, cell.column + 1).value
                                elif cell.row + 1 <= max_row:
                                    next_val = sheet.cell(cell.row + 1, cell.column).value
                            except:
                                pass
                            extracted[key] = clean_value(next_val, key, cell)

    # Ensure CUSTOMER field is handled properly
    if 'CUSTOMER' not in extracted:
        print("CUSTOMER field not found, checking again...")
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                cell_text = str(sheet.cell_value(row_idx, col_idx)).strip().lower()
                if re.search(r"(customer|client)", cell_text, re.IGNORECASE):
                    customer_data = clean_value(sheet.cell_value(row_idx, col_idx), "CUSTOMER")
                    if customer_data:
                        extracted['CUSTOMER'] = customer_data
                        print(f"Found CUSTOMER: {customer_data}")
                        break
    return extracted


# Write to final Excel without formatting
with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
    for subfolder in os.listdir(main_folder_path):
        subfolder_path = os.path.join(main_folder_path, subfolder)
        if os.path.isdir(subfolder_path):
            data = []
            for filename in os.listdir(subfolder_path):
                if filename.endswith((".xls", ".xlsx")):
                    file_path = os.path.join(subfolder_path, filename)
                    row_data = {"File Name": filename}
                    try:
                        if filename.endswith(".xls"):
                            book = xlrd.open_workbook(file_path)
                            sheet = book.sheet_by_index(0)
                            extracted = extract_from_xls(sheet)
                        else:
                            wb = openpyxl.load_workbook(file_path, data_only=True)
                            sheet = wb.active
                            extracted = extract_from_xlsx(sheet)
                        row_data.update(extracted)
                        data.append(row_data)
                    except Exception as e:
                        print(f"❌ Error reading {filename} in {subfolder}: {e}")

            if data:
                df = pd.DataFrame(data)
                sheet_name = subfolder[:31]
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                worksheet = writer.sheets[sheet_name]
                workbook = writer.book  # needed for formatting

                # Set column width
                for col_num in range(len(df.columns)):
                    worksheet.set_column(col_num, col_num, 25)

                # ✅ Apply conditional formatting: highlight empty cells
                # This will highlight any blank cell with light red, but remove the color automatically if user fills it
                format_blank = workbook.add_format({'bg_color': '#FFC7CE'})
                worksheet.conditional_format(1, 0, len(df), len(df.columns) - 1, {
                    'type': 'blanks',
                    'format': format_blank
                })


print(f"✅ All summaries saved to {output_file}")
