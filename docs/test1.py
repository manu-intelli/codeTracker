import os
import re
import xlrd
import openpyxl
from datetime import datetime
from openpyxl.cell.cell import Cell
import pandas as pd

# MAIN folder path
main_folder_path = r"D:\RIghtstroken\Pricing\PricingExtraction\Pricing_Sheets_24-Feb-25"
output_file = "final_summary_all_folders.xlsx"

# Regex patterns
patterns = {
    "Part# / Model name": r"(part\s*#|model\s*name)",
    "OPP#": r"opp\s*#?",
    "CUSTOMER": r"customer",
    "Assembly cost / PPD": r"\b(assembly cost|ppd)\b",
    "Estimated BOM cost": r"\b(estimated bom cost|bom cost per unit)\b",
    "Design & Development cost": r"design and development cost",
    "Recommended Price": r"recommended price",
    "Comments to Steven": r"(comments for steven\.s|comments to steven)",
    "CREATED ON": r"created\s*on\s*[:\-]?"
}

# Clean values (handle Excel serial date numbers)
def clean_value(value, key, cell=None):
    if isinstance(cell, Cell) and cell.is_date:
        return cell.value.strftime("%m/%d/%Y")
    if isinstance(value, float) and key == "CREATED ON":
        try:
            # Assume it's an Excel serial date from .xls
            date_value = datetime(*xlrd.xldate_as_tuple(value, 0))
            return date_value.strftime("%m/%d/%Y")
        except:
            pass
    return str(value).strip()

# Read .xls
def extract_from_xls(sheet):
    extracted = {}
    for row_idx in range(sheet.nrows):
        for col_idx in range(sheet.ncols):
            cell_value = str(sheet.cell_value(row_idx, col_idx)).strip().lower()
            for key, pattern in patterns.items():
                if re.search(pattern, cell_value, re.IGNORECASE):
                    try:
                        next_value = sheet.cell_value(row_idx, col_idx + 1)
                        if key not in extracted:
                            extracted[key] = clean_value(next_value, key)
                    except:
                        continue
    return extracted

# Read .xlsx
def extract_from_xlsx(sheet):
    extracted = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                value = str(cell.value).strip().lower()
                for key, pattern in patterns.items():
                    if re.search(pattern, value, re.IGNORECASE):
                        try:
                            next_cell = sheet.cell(cell.row, cell.column + 1)
                            if key not in extracted:
                                extracted[key] = clean_value(next_cell.value, key, next_cell)
                        except:
                            continue
    return extracted

# Write to Excel with one sheet per subfolder
with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
    for subfolder in os.listdir(main_folder_path):
        subfolder_path = os.path.join(main_folder_path, subfolder)
        if os.path.isdir(subfolder_path):
            data = []
            for filename in os.listdir(subfolder_path):
                if filename.endswith(".xls") or filename.endswith(".xlsx"):
                    file_path = os.path.join(subfolder_path, filename)
                    row_data = {"File Name": filename}
                    try:
                        if filename.endswith(".xls"):
                            book = xlrd.open_workbook(file_path)
                            sheet = book.sheet_by_index(0)
                            extracted = extract_from_xls(sheet)
                        else:
                            wb = openpyxl.load_workbook(file_path, data_only=True)
                            sheet = wb.active
                            extracted = extract_from_xlsx(sheet)
                        row_data.update(extracted)
                        data.append(row_data)
                    except Exception as e:
                        print(f"❌ Error reading {filename} in {subfolder}: {e}")
            if data:
                df = pd.DataFrame(data)
                sheet_name = subfolder[:31]  # Excel allows max 31 chars for sheet name
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                worksheet = writer.sheets[sheet_name]
                workbook = writer.book
                text_format = workbook.add_format({'num_format': '@'})
                for col_num in range(len(df.columns)):
                    worksheet.set_column(col_num, col_num, 25, text_format)

print(f"✅ All summaries saved to {output_file}")
